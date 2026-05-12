"""
Data ingestion script.
Reads Excel files and loads enterprise architecture data into Firestore.
Run this once to populate the knowledge base.

Usage:
    python -m scripts.ingest_data
"""
import os
import sys
import openpyxl
from google.cloud import firestore

# Configuration
GCP_PROJECT_ID = "roger-496113"
BASE_PATH = os.environ.get(
    "DATA_PATH",
    r"C:\Users\ASUS\Documents\Ideam\Arquitectura\Mapa_Informacion_Institucional",
)

db = firestore.Client(project=GCP_PROJECT_ID)


def clean_value(val):
    """Clean a cell value for storage."""
    if val is None:
        return ""
    return str(val).strip()


def ingest_datos_maestros():
    """
    Ingest master data tables from Formato_datos_maestros_referencia_diligenciado.xlsx.
    Creates documents in the 'datos_maestros' collection.
    """
    print("📊 Ingesting datos maestros...")
    filepath = os.path.join(BASE_PATH, "Formato_datos_maestros_referencia_diligenciado.xlsx")
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Hoja1"]

    # Headers are in row 2
    count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        consecutivo = clean_value(row[0])
        if not consecutivo:
            continue

        nombre_tabla = clean_value(row[2])
        # Create a clean document ID from the table name
        doc_id = nombre_tabla.upper().replace(" ", "_").replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U").replace("Ñ", "N")

        atributos_str = clean_value(row[5])
        atributos = [a.strip() for a in atributos_str.split(",") if a.strip()]

        doc = {
            "consecutivo": int(consecutivo) if consecutivo.isdigit() else consecutivo,
            "area": clean_value(row[1]),
            "nombre_tabla": nombre_tabla,
            "descripcion": clean_value(row[3]),
            "cantidad_atributos": clean_value(row[4]),
            "atributos": atributos,
            "cantidad_registros": clean_value(row[6]),
            "usuario_interno": clean_value(row[7]),
            "frecuencia_actualizacion": clean_value(row[8]),
            "aprovechamiento": clean_value(row[9]),
            "georreferenciado": clean_value(row[10]),
            "atributos_georeferenciados": clean_value(row[11]),
            "productor_consumidor": clean_value(row[12]),
            "medio_intercambio": clean_value(row[13]),
            "almacena_en_bd": clean_value(row[14]),
            "entidad_origen": clean_value(row[15]),
        }

        db.collection("datos_maestros").document(doc_id).set(doc)
        count += 1
        print(f"  ✅ {nombre_tabla} ({doc_id})")

    print(f"  📊 Total: {count} tablas ingested\n")


def ingest_diccionario():
    """
    Ingest data dictionary from gci-oe-f014_formato_diccionario_de_datos_0.xlsx.
    Creates documents in the 'diccionario_datos' collection.
    """
    print("📋 Ingesting diccionario de datos...")
    filepath = os.path.join(BASE_PATH, "gci-oe-f014_formato_diccionario_de_datos_0.xlsx")
    wb = openpyxl.load_workbook(filepath)
    ws = wb["DICCIONARIO DE DATOS"]

    # Headers are in row 6-7, data starts at row 8
    count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=8, values_only=True)):
        nombre_archivo = clean_value(row[1])
        nombre_variable = clean_value(row[3])

        if not nombre_archivo or not nombre_variable:
            continue

        # Map variable to a master table reference
        tabla_ref = _map_archivo_to_tabla(nombre_archivo)

        doc_id = f"{nombre_archivo}_{nombre_variable}".replace(".", "_").replace(" ", "_")[:100]

        llave_primaria_val = clean_value(row[5])
        llave_foranea_val = clean_value(row[6])

        doc = {
            "nombre_archivo": nombre_archivo,
            "descripcion_archivo": clean_value(row[2]),
            "nombre_variable": nombre_variable,
            "descripcion_variable": clean_value(row[4]),
            "llave_primaria": llave_primaria_val.upper() in ("SÍ", "SI", "S", "X", "YES"),
            "llave_foranea": llave_foranea_val.upper() in ("SÍ", "SI", "S", "X", "YES"),
            "condicion": clean_value(row[7]),
            "tipo_dato": clean_value(row[8]),
            "longitud": clean_value(row[9]),
            "dominios": clean_value(row[10]),
            "regla_validacion": clean_value(row[11]),
            "observaciones": clean_value(row[12]),
            "tabla_referencia": tabla_ref,
        }

        db.collection("diccionario_datos").document(doc_id).set(doc)
        count += 1

    print(f"  📋 Total: {count} variables ingested\n")


def _map_archivo_to_tabla(nombre_archivo):
    """
    Map a file name from the dictionary to its corresponding master table.
    Based on the descriptions and relationships in the data.
    """
    nombre_lower = nombre_archivo.lower()

    mapping = {
        "aut_prec": "PRECIPITACION_ESTACIONES",
        "precipitacion": "PRECIPITACION_ESTACIONES",
        "geojson": "GEOGRAFIA",
        "departamento": "GEOGRAFIA",
        "municipio": "GEOGRAFIA",
        "pais": "GEOGRAFIA",
        "rio": "GEOGRAFIA",
        "szh": "GEOGRAFIA",
        "categorias_amenaza": "CATALOGO_AMENAZA",
        "colores_leyenda": "VISUALIZACION",
        "configuracion_mapas": "CONFIG_MAPA",
        "limites_mapas": "CONFIG_MAPA",
        "layout_": "CONFIG_MAPA",
        "alerta": "ALERTAS_IDD",
        "amenaza_idd": "AMENAZA_IDD",
        "probabilidad_idd": "AMENAZA_IDD",
        "deslizamiento": "AMENAZA_IDD",
        "seguimiento": "AMENAZA_IDD",
        "municipios_eliminados": "CONTROL_CALIDAD",
        "nombresvariables": "CONTROL_CALIDAD",
    }

    for key, tabla in mapping.items():
        if key in nombre_lower:
            return tabla

    return "GENERAL"


def ingest_flujos():
    """
    Ingest information flow descriptions based on the ArchiMate diagrams.
    Creates documents in the 'flujos_informacion' collection.
    """
    print("🔄 Ingesting flujos de información...")

    flujos = [
        {
            "nombre": "Flujo Simplificado - Precipitación (Nivel 0)",
            "descripcion": (
                "Flujo de alto nivel del proceso de precipitación. "
                "Los usuarios y observadores voluntarios alimentan el sistema "
                "a través de BARFTP y módulos personalizados hacia Polaris, "
                "Aquarius Time Series y Cassandra. El procesamiento incluye "
                "descargue y consolidación de datos de precipitación, "
                "culminando en la generación de mapas."
            ),
            "etapas": [
                "Aprovisionamiento (Usuarios, BARFTP, Observadores)",
                "Polaris / Aquarius / Cassandra",
                "Descargue y consolidación",
                "Generación de mapas",
            ],
            "sistemas": ["Polaris", "Aquarius Time Series", "Cassandra", "BARFTP"],
            "imagen": "img_0_9_3.png",
            "imagen_url": "",
            "tipo": "nivel_0",
        },
        {
            "nombre": "Flujo Simplificado - Precipitación",
            "descripcion": (
                "Flujo simplificado que muestra el ciclo completo de datos "
                "de precipitación desde el aprovisionamiento hasta la distribución. "
                "Usuarios → BARFTP → Polaris/Aquarius/Cassandra → "
                "Descargue y consolidación → Ejecución de modelos de predicción → "
                "Elaboración de informe → Generación de mapas."
            ),
            "etapas": [
                "Aprovisionamiento",
                "Descargue y consolidación",
                "Ejecución de modelos de predicción",
                "Elaboración de informe",
                "Generación de mapas",
            ],
            "sistemas": ["Polaris", "Aquarius Time Series", "Cassandra", "BARFTP"],
            "imagen": "img_0_2_4.png",
            "imagen_url": "",
            "tipo": "simplificado",
        },
        {
            "nombre": "Flujo Detallado - Precipitación e IDD (Deslizamientos)",
            "descripcion": (
                "Flujo detallado que integra precipitación con el modelo IDD "
                "(Índice Detonante de Deslizamientos). Incluye: estaciones "
                "convencionales, observadores voluntarios, módulo personalizado, "
                "archivos de configuración (mapas, leyendas, shapefiles), "
                "cálculo de precipitación acumulada, pronóstico GFS, "
                "modelo IDD, interpolación IDW, generación de mapas de alerta "
                "y seguimiento de deslizamientos. Distribuye informes de "
                "amenaza por deslizamiento de tierra."
            ),
            "etapas": [
                "Aprovisionamiento (múltiples fuentes)",
                "Descargue y consolidación",
                "Cálculo precipitación acumulada",
                "Pronóstico GFS",
                "Modelo IDD",
                "Interpolación IDW",
                "Generación de mapas (MapaAlertas_IDD.py)",
                "Seguimiento (Actualiza_Seguimiento_s_Deslizamientos.py)",
                "Informe de amenaza",
            ],
            "sistemas": [
                "Polaris", "Aquarius Time Series", "Cassandra",
                "BARFTP", "Q-Connect", "NOAA", "GFS",
            ],
            "imagen": "img_0_3_3.png",
            "imagen_url": "",
            "tipo": "detallado",
        },
        {
            "nombre": "Flujo - Cálculo de Emisiones GEI (SINGEI)",
            "descripcion": (
                "Flujo del Sistema Nacional de Inventario de Gases de Efecto "
                "Invernadero. Desde MariaDB se extraen datos de actividad (B0), "
                "factores de emisión y parámetros IPCC. El procesamiento incluye: "
                "selección y extracción filtrada por año/vigencia/categoría IPCC/sector, "
                "integración de datos de actividad y factores, ejecución de cálculo "
                "de emisiones con fórmulas IPCC, validación de resultados, "
                "agregación por sector/categoría/período, y transformación ETL. "
                "Los resultados se almacenan en base de datos analítica con "
                "emisiones GEI, indicadores y series históricas."
            ),
            "etapas": [
                "Almacenamiento MariaDB",
                "Selección y extracción",
                "Integración datos + factores",
                "Cálculo emisiones (IPCC)",
                "Validación",
                "Agregación y consolidación",
                "Transformación ETL",
                "Base analítica",
                "Consumo de información",
            ],
            "sistemas": ["MariaDB"],
            "imagen": "img_0_5_3.png",
            "imagen_url": "",
            "tipo": "singei",
        },
        {
            "nombre": "Flujo - Registro de Datos de Actividad",
            "descripcion": (
                "Flujo de ingesta de datos de actividad para el cálculo de "
                "emisiones GEI. Los usuarios registran: valores de consumo/producción, "
                "unidades de medida, año, ubicación geográfica y categoría IPCC. "
                "El procesamiento incluye validación (tipos, obligatoriedad, rangos), "
                "estandarización (unidades, formatos, codificación), enriquecimiento "
                "(id_categoria, id_sector, relaciones) y asignación de estado. "
                "Los datos se almacenan en MariaDB y se disponibilizan para "
                "procesos analíticos y cálculo de emisiones GEI."
            ),
            "etapas": [
                "Registro de datos de actividad",
                "Procesamiento y control de calidad",
                "Almacenamiento MariaDB",
                "Disponibilización para procesos analíticos",
                "Cálculo de emisiones GEI",
            ],
            "sistemas": ["MariaDB"],
            "imagen": "img_0_6_3.png",
            "imagen_url": "",
            "tipo": "datos_actividad",
        },
        {
            "nombre": "Flujo - Factores de Emisión",
            "descripcion": (
                "Flujo de gestión de factores de emisión para el inventario "
                "de GEI. Los usuarios registran: factores de emisión, "
                "contaminantes (CO2, CH4, N2O), unidades de medida, "
                "nivel o metodología IPCC, y vigencia/año. El procesamiento "
                "incluye validación (campos obligatorios, rangos válidos, "
                "consistencia de unidades, unicidad por combinación), "
                "estandarización y normalización (codificación IPCC, "
                "homologación de unidades, asociación a catálogos). "
                "Se almacena en MariaDB (tn_singei_f_dos, tablas paramétricas). "
                "Distribución: preparación para cálculo analítico, "
                "exportación en excel/csv, y consulta operativa."
            ),
            "etapas": [
                "Registro y edición de factores",
                "Validación y control",
                "Estandarización IPCC",
                "Almacenamiento MariaDB",
                "Preparación para cálculo",
                "Cálculo de emisiones",
                "Exportación (excel/csv)",
                "Consulta operativa",
            ],
            "sistemas": ["MariaDB"],
            "imagen": "img_0_7_3.png",
            "imagen_url": "",
            "tipo": "factores_emision",
        },
    ]

    for flujo in flujos:
        doc_id = flujo["nombre"][:50].replace(" ", "_").replace("-", "").replace("(", "").replace(")", "")
        db.collection("flujos_informacion").document(doc_id).set(flujo)
        print(f"  ✅ {flujo['nombre']}")

    print(f"  🔄 Total: {len(flujos)} flujos ingested\n")


def ingest_metadata():
    """Ingest general metadata about the knowledge base."""
    print("📝 Ingesting metadata...")

    metadata = {
        "institucion": "IDEAM",
        "area_principal": "OSPA - Oficina del Servicio de Pronósticos y Alertas",
        "marco_normativo": "SEN - Sistema Estadístico Nacional",
        "formato_referencia": "GCI-OE-F014",
        "version_formato": "01",
        "fecha_formato": "16/02/2026",
        "fuente_referencia": "https://www.sen.gov.co",
        "descripcion": (
            "Mapa de Información Institucional del IDEAM. "
            "Incluye datos maestros y de referencia, diccionario de datos "
            "institucional, y flujos de información para precipitación, "
            "alertas por deslizamientos (IDD), y emisiones de gases de "
            "efecto invernadero (SINGEI)."
        ),
        "areas_tematicas": [
            "Precipitación y meteorología",
            "Amenazas por deslizamientos (IDD)",
            "Emisiones GEI (SINGEI)",
            "Geografía y capas espaciales",
        ],
    }

    db.collection("metadata").document("general").set(metadata)
    print("  ✅ Metadata general ingested\n")


def main():
    """Run the full ingestion pipeline."""
    print("🚀 Starting data ingestion for Roger...\n")
    print(f"   Project: {GCP_PROJECT_ID}")
    print(f"   Data path: {BASE_PATH}\n")

    ingest_datos_maestros()
    ingest_diccionario()
    ingest_flujos()
    ingest_metadata()

    print("✅ Data ingestion complete!")


if __name__ == "__main__":
    main()
